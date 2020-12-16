from openpyxl import Workbook,load_workbook

wb = Workbook()
ws = wb.active

ws.title = 'kerjabaru'

ws1 = wb.create_sheet("kerja")
ws2 = wb.create_sheet("kerja", 0)
ws3 = wb.create_sheet("kerja", -1)

print(wb.sheetnames)

workaktif = wb['kerjabaru']
print (workaktif)


a = ['aku', 'suka', 'makan', 'bakso']

for row in range(0,len(a)):
    ws.cell(row=row+1, column=1).value= a[row]

for column in range(0,len(a)):
    ws.cell(row=1, column=column+1).value= a[column]

menu = [['hari'],['senin', 'nasi', 'ayam'],['selasa', 'susu'],['rabu','nasi goreng', 'ati ampela', 'jus apel'],['kamis','capcay','telur mata sapi']]
for item in menu:
    ws.append(item)

wb.save('namafile.xlsx')
